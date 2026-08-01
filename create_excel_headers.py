from pathlib import Path
from openpyxl import Workbook, load_workbook

path = Path(r"c:\Users\user\OneDrive\Documents\GitHub\OrangeHRM-Manual-Testing\Login Module\02_Test_Scenarios.xlsx")
headers = ["Scenario ID", "Module", "Test Scenario", "Priority", "Type", "Status", "Remarks"]

if path.exists():
    wb = load_workbook(path)
else:
    wb = Workbook()

ws = wb.active
ws.title = "Test Scenarios"

for col_idx, value in enumerate(headers, start=1):
    ws.cell(row=1, column=col_idx, value=value)

for col_idx in range(1, len(headers) + 1):
    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20

wb.save(path)
print(f"Updated {path}")
print(headers)
